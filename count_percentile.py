import pandas as pd
import os
import glob
import math
import openpyxl
import re
import numpy as np
from scipy.stats import percentileofscore

def is_url(text):
    """Checks if the text is a URL"""
    if not isinstance(text, str):
        return False
    return text.startswith(('http://', 'https://', 'www.'))

def parse_score_value(score_text):
    """
    Parses score value, handling various formats:
    - Single numbers: "0.85" -> 0.85
    - Multiple numbers separated by comma: "0.85, 0.90" -> 0.85 (first value)
    - URL: "https://..." -> None
    - Non-numeric values: "N/A" -> None

    Returns:
        float or None: Parsed numeric value or None
    """
    if pd.isna(score_text):
        return None

    # Convert to string
    score_str = str(score_text).strip()

    # Skip URLs
    if is_url(score_str):
        return None

    # Skip empty strings
    if not score_str or score_str.lower() in ['', 'n/a', 'na', 'null', 'none']:
        return None

    # If there are commas, take the first value
    if ',' in score_str:
        # Split by comma and take the first value
        parts = score_str.split(',')
        first_part = parts[0].strip()

        # Try to convert to number
        try:
            return float(first_part)
        except ValueError:
            return None

    # Try to convert to number
    try:
        return float(score_str)
    except ValueError:
        return None

def convert_scores_to_percentiles(input_xlsx_path, leaderboards_folder, output_xlsx_path):
    """
    Converts model results to percentiles based on competition leaderboards

    Args:
        input_xlsx_path (str): Path to input Excel file with results
        leaderboards_folder (str): Folder with CSV files of leaderboards
        output_xlsx_path (str): Path to output Excel file
    """

    # Read Excel file, only SCORES sheet
    print(f"Reading file: {input_xlsx_path}")
    df = pd.read_excel(input_xlsx_path, sheet_name='SCORES', header=None)

    # Create a copy for processing
    result_df = df.copy()

    # Get list of all CSV files in the leaderboards folder
    csv_files = glob.glob(os.path.join(leaderboards_folder, "*.csv"))
    leaderboards = {}

    print(f"Found {len(csv_files)} leaderboard files:")

    # Load all leaderboards
    for csv_file in csv_files:
        # Filename without extension - this is the competition name
        competition_name = os.path.splitext(os.path.basename(csv_file))[0]
        print(f"  - {competition_name}")

        try:
            # Read leaderboard
            leaderboard_df = pd.read_csv(csv_file)

            # Check for Rank column
            if 'Rank' not in leaderboard_df.columns:
                print(f"    Warning: file {csv_file} does not have 'Rank' column")
                continue

            # Check for Score column
            if 'Score' not in leaderboard_df.columns:
                print(f"    Warning: file {csv_file} does not have 'Score' column")
                continue

            # Remove rows without rank or score
            leaderboard_df = leaderboard_df.dropna(subset=['Rank'])

            # Convert Score to numeric values, remove non-numeric
            leaderboard_df['Score'] = pd.to_numeric(leaderboard_df['Score'], errors='coerce')
            leaderboard_df = leaderboard_df.dropna(subset=['Score'])

            total_teams = len(leaderboard_df)

            if total_teams == 0:
                print(f"    Warning: file {csv_file} does not contain valid data")
                continue

            leaderboards[competition_name] = {
                'df': leaderboard_df,
                'total_teams': total_teams
            }
            print(f"    Loaded leaderboard with {total_teams} teams")

        except Exception as e:
            print(f"    Error reading {csv_file}: {e}")

    print(f"\nSuccessfully loaded {len(leaderboards)} leaderboards")

    # Get result columns (K to CJ inclusive, indices 10-87)
    # In pandas column indexing starts from 0
    start_col = 11  # Column K
    end_col = min(88, len(df.columns) - 1)  # Column CJ or last available

    print(f"\nProcessing columns with indices {start_col} to {end_col}")

    # Update counter
    updates_count = 0

    # Process all data rows (starting from index 2, as first 2 rows are headers)
    for idx in range(2, len(df)):  # Start from index 2 (3rd row in Excel)
        row = df.iloc[idx]
        competition = row.iloc[4]  # competition is in column E (index 4)

        if pd.isna(competition):
            continue

        # Convert datetime objects to strings to match leaderboard file names
        if isinstance(competition, pd.Timestamp):
            # Convert datetime to YYYY-MM-DD format
            competition_str = competition.strftime('%Y-%m-%d')
        elif hasattr(competition, 'strftime'):
            # Handle other datetime object types
            competition_str = competition.strftime('%Y-%m-%d')
        else:
            competition_str = str(competition)

        # Check if there is a leaderboard for this competition
        if competition_str not in leaderboards:
            print(f"    Warning: leaderboard for competition '{competition_str}' not found")
            continue

        total_teams = leaderboards[competition_str]['total_teams']
        comp_name = row.iloc[1] if not pd.isna(row.iloc[1]) else 'Unknown'  # comp_name in column B (index 1)
        print(f"\nProcessing row {idx + 1}: {comp_name} ({competition_str})")

        # Process result columns
        for col_idx in range(start_col, end_col + 1):
            if col_idx >= len(df.columns):
                break

            score_raw = row.iloc[col_idx]

            # Parse score value
            score = parse_score_value(score_raw)

            # Skip if parsing failed
            if score is None:
                # Show only first few problematic values for debugging
                if updates_count < 5 and not pd.isna(score_raw):
                    print(f"    Skipped value: '{score_raw}' (type: {type(score_raw)})")
                continue

            # Find percentile for this score
            percentile = calculate_percentile(score, leaderboards[competition_str]['df'], total_teams)

            if percentile is not None:
                result_df.iloc[idx, col_idx] = percentile
                updates_count += 1

    print(f"\nTotal updated values: {updates_count}")

    # Save result preserving structure
    print(f"Saving result to: {output_xlsx_path}")

    # Load original file to preserve structure
    original_wb = openpyxl.load_workbook(input_xlsx_path)
    original_ws = original_wb['SCORES']

    # Create new file with same structure
    with pd.ExcelWriter(output_xlsx_path, engine='openpyxl') as writer:
        # Create new sheet
        new_ws = writer.book.create_sheet('SCORES')

        # Copy first two rows (headers) from original file
        for row_num in range(1, 3):  # Copy rows 1 and 2
            for col_num in range(1, original_ws.max_column + 1):
                original_cell = original_ws.cell(row=row_num, column=col_num)
                new_cell = new_ws.cell(row=row_num, column=col_num)
                new_cell.value = original_cell.value

        # Write data starting from 3rd Excel row (index 2 in pandas)
        for idx in range(2, len(result_df)):  # Start from index 2 (3rd row in Excel)
            row = result_df.iloc[idx]
            excel_row = idx + 1  # +1 because Excel indexing starts from 1, pandas from 0
            for col_idx, value in enumerate(row, 1):
                new_ws.cell(row=excel_row, column=col_idx, value=value)

        # Remove default sheet
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])

    print("Done!")

    return result_df

def calculate_percentile(our_score, leaderboard_df, total_teams):
    """
    Calculates percentile for our score based on leaderboard using np.percentile

    Args:
        our_score (float): Our score
        leaderboard_df (DataFrame): Leaderboard data
        total_teams (int): Total number of teams

    Returns:
        int or None: Percentile (integer) or None if calculation failed
    """
    try:
        # Get all scores from leaderboard and convert to numeric
        scores_series = pd.to_numeric(leaderboard_df['Score'], errors='coerce')
        scores = scores_series.dropna().values

        if len(scores) == 0:
            print(f"    Warning: no valid numeric scores in leaderboard")
            return None

        # our_score should already be float after parsing
        if not isinstance(our_score, (int, float)):
            print(f"    Warning: unexpected score type: {type(our_score)}")
            return None

        # Determine sorting direction by ranks
        # Sort leaderboard by ranks and look at trend
        leaderboard_sorted = leaderboard_df.sort_values('Rank')
        first_scores = pd.to_numeric(leaderboard_sorted['Score'].head(10), errors='coerce').dropna()
        last_scores = pd.to_numeric(leaderboard_sorted['Score'].tail(10), errors='coerce').dropna()

        # If top places have higher Score, then higher is better
        higher_is_better = first_scores.mean() > last_scores.mean() if len(first_scores) > 0 and len(last_scores) > 0 else True

        # Correct percentile calculation using scipy.stats.percentileofscore:
        # percentileofscore finds percentage of scores that are less than or equal to ours
        
        if higher_is_better:
            # For higher_is_better: use percentileofscore to find rank
            # percentileofscore returns percentage of scores <= ours
            # We need percentage of scores < ours (worse than ours)
            percentile = percentileofscore(scores, our_score, kind='rank') - 1
            # Ensure result is not negative
            percentile = max(0, percentile)
        else:
            # For lower_is_better: invert logic
            # percentileofscore returns percentage of scores <= ours
            # For lower_is_better we need percentage of scores > ours (worse than ours)
            percentile = 100 - percentileofscore(scores, our_score, kind='rank')

        # Round to integer
        result = round(percentile)

        # Limit range to 1-100
        result = max(1, min(100, result))

        return result

    except Exception as e:
        print(f"    Error calculating percentile for score {our_score}: {e}")
        return None


def main(input_file, output_file):
    # Usage example
    leaderboards_dir = "leaderboards"  # folder with CSV files of leaderboards
    convert_scores_to_percentiles(input_file, leaderboards_dir, output_file)

if __name__ == "__main__":
    input_file = "final_version_tmp.xlsx"
    output_file = "final_version_percentiles.xlsx"
    main(input_file)